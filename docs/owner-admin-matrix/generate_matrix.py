"""Generate owner/admin access matrix Excel files for Sojori Orchestrator vs Dashboard."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parent

HEADER_FILL = PatternFill("solid", fgColor="B8851A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="CCCCCC")

COLUMNS = [
    ("Page / sous-page", 28),
    ("Route (nouveau)", 32),
    ("Accès recommandé", 22),
    ("Équivalent ancien dashboard", 36),
    ("Comportement ancien (admin)", 42),
    ("Comportement ancien (owner)", 32),
    ("État nouveau front", 18),
    ("Filtre owner admin — pattern", 28),
    ("Comportement cible admin", 40),
    ("Comportement cible owner", 28),
    ("Priorité impl.", 12),
]


def style_sheet(ws, row_count: int):
    for col_idx, (_, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(2, row_count + 2):
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_workbook(filename: str, tab_name: str, rows: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = tab_name[:31]
    headers = [h for h, _ in COLUMNS]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(k, "") for k, _ in COLUMNS])
    style_sheet(ws, len(rows))
    path = OUT_DIR / filename
    wb.save(path)
    return path


# Shared pattern labels
P_ALL_DEFAULT = "Barre Propriétaire : « Tous » = données agrégées tous owners ; sélection = filterOwnerId API"
P_BLOCK_OWNER = "Chips Admin / Owner : scope Owner → bloquer chargement tant qu'owner non choisi"
P_COLUMN_FILTER = "Colonne Owner visible admin + filtre toolbar (Tous par défaut)"
P_ORCH_CHIPS = "Chips Admin config / Owner config + select owner (config uniquement)"
P_NONE = "Pas de filtre UI ; scope JWT backend uniquement"
P_MOCK_ONLY = "UI mock — pas de RBAC ni filtre branchés"

ACCESS_ALL = "Admin + Owner + Worker"
ACCESS_ADMIN = "Admin + SuperAdmin uniquement"
ACCESS_ADMIN_OWNER = "Admin + Owner (+ Worker selon grants)"

SHEETS = {
    "01_Pilotage.xlsx": {
        "tab": "Pilotage",
        "rows": [
            {
                "Page / sous-page": "Dashboard",
                "Route (nouveau)": "/dashboard",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Overview → BusinessOverview (tab new)",
                "Comportement ancien (admin)": "Pas de barre owner globale sur Overview ; KPIs via API selon JWT (souvent tous les owners)",
                "Comportement ancien (owner)": "Données limitées à son compte (JWT)",
                "État nouveau front": "Partiel (API + mock fallback)",
                "Filtre owner admin — pattern": P_NONE + " — optionnel : sélecteur owner avant KPIs si agrégats cross-tenant",
                "Comportement cible admin": "Recommandé : ne pas bloquer l'écran ; filtre owner optionnel en tête (comme Financials ancien) OU agrégat global par défaut",
                "Comportement cible owner": "Vue auto-scopée, pas de filtre owner",
                "Priorité impl.": "P1",
            },
            {
                "Page / sous-page": "Analytics",
                "Route (nouveau)": "/analytics",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Report → Financials (financial/Analytics)",
                "Comportement ancien (admin)": P_ALL_DEFAULT,
                "Comportement ancien (owner)": "Données scopées compte",
                "État nouveau front": "Partiel (pas de filtre owner)",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Afficher analytics tous owners par défaut ; filtre owner restreint les séries",
                "Comportement cible owner": "Ses listings / réservations uniquement",
                "Priorité impl.": "P1",
            },
            {
                "Page / sous-page": "Reports",
                "Route (nouveau)": "/reports",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Report → Rapports (4 tabs) + PM monthly",
                "Comportement ancien (admin)": "AdminOwnerScopeLayout : filtre owner inline ; rapports avec colonne ownerName",
                "Comportement ancien (owner)": "Rapports de son périmètre",
                "État nouveau front": P_MOCK_ONLY,
                "Filtre owner admin — pattern": P_COLUMN_FILTER,
                "Comportement cible admin": "Colonne Owner dans historique + filtre propriétaire ; génération rapport filtrée si owner sélectionné",
                "Comportement cible owner": "Pas de colonne/filtre owner",
                "Priorité impl.": "P2",
            },
            {
                "Page / sous-page": "Orchestration (entrée sidebar)",
                "Route (nouveau)": "/orchestration",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Orchestration → Chronologie",
                "Comportement ancien (admin)": "Liste/timeline toutes réservations (scope JWT) ; pas de barre owner dédiée sur timeline",
                "Comportement ancien (owner)": "Ses réservations / plans",
                "État nouveau front": "Mock UI",
                "Filtre owner admin — pattern": "Filtre owner optionnel (listing/résa) — pas bloquant",
                "Comportement cible admin": "Voir toutes les chronologies ; filtre owner optionnel",
                "Comportement cible owner": "Chronologies de son parc",
                "Priorité impl.": "P1",
            },
        ],
    },
    "02_Orchestration_sous_pages.xlsx": {
        "tab": "Orchestration détail",
        "rows": [
            {
                "Page / sous-page": "Plans d'orchestration",
                "Route (nouveau)": "/orchestration/plans, /admin/orchestrator",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "(plans — API srv-orchestrator)",
                "Comportement ancien (admin)": "Accès API cross-owner",
                "Comportement ancien (owner)": "Plans de ses réservations",
                "État nouveau front": "API branchée, pas de filtre owner UI",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Liste tous plans ; filtre owner réduit la liste",
                "Comportement cible owner": "Auto-filtré backend",
                "Priorité impl.": "P1",
            },
            {
                "Page / sous-page": "Timeline réservation",
                "Route (nouveau)": "/orchestration/timeline/:id",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Chronologie détail réservation",
                "Comportement ancien (admin)": "Accès si réservation visible (tous owners)",
                "Comportement ancien (owner)": "Ses réservations uniquement",
                "État nouveau front": "Partiel",
                "Filtre owner admin — pattern": "N/A (page détail)",
                "Comportement cible admin": "Accès lecture/ actions admin sur toute résa",
                "Comportement cible owner": "Lecture/ actions sur ses résas",
                "Priorité impl.": "P1",
            },
            {
                "Page / sous-page": "Événements système",
                "Route (nouveau)": "/orchestration/events",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Orchestration → Event tab",
                "Comportement ancien (admin)": "Log global événements",
                "Comportement ancien (owner)": "Événements de son parc",
                "État nouveau front": "Partiel / mock",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Tous événements + filtre owner",
                "Comportement cible owner": "Scope JWT",
                "Priorité impl.": "P2",
            },
            {
                "Page / sous-page": "Configuration workflows",
                "Route (nouveau)": "/orchestration/config",
                "Accès recommandé": ACCESS_ADMIN_OWNER,
                "Équivalent ancien dashboard": "Orchestration → Configuration (OrchestratorConfigContent)",
                "Comportement ancien (admin)": P_ORCH_CHIPS + " ; Owner scope : block si orchOwnerId vide",
                "Comportement ancien (owner)": "Édition de sa config owner (pas les chips admin)",
                "État nouveau front": "Non intégré",
                "Filtre owner admin — pattern": P_BLOCK_OWNER,
                "Comportement cible admin": "Chip Admin config = template plateforme ; Chip Owner config = choisir owner puis éditer ; pas de données owner sans sélection",
                "Comportement cible owner": "Directement sa config, pas de sélecteur",
                "Priorité impl.": "P0",
            },
        ],
    },
    "03_Calendrier.xlsx": {
        "tab": "Calendrier",
        "rows": [
            {
                "Page / sous-page": "Vue multi-propriétés (Inventory)",
                "Route (nouveau)": "/calendar",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Calendar → Inventory",
                "Comportement ancien (admin)": "Calendrier multi-listings ; scope JWT (souvent tous) ; pas AdminOwnerScopeLayout sur calendar",
                "Comportement ancien (owner)": "Ses listings",
                "État nouveau front": "Partiel",
                "Filtre owner admin — pattern": "Filtre owner recommandé (liste déroulante) — défaut = tous",
                "Comportement cible admin": "Voir tous les logements ; filtre owner restreint les lignes du Gantt",
                "Comportement cible owner": "Uniquement ses propriétés",
                "Priorité impl.": "P1",
            },
        ],
    },
    "04_Reservations.xlsx": {
        "tab": "Réservations",
        "rows": [
            {
                "Page / sous-page": "Liste réservations",
                "Route (nouveau)": "/reservations",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Reservation → Liste",
                "Comportement ancien (admin)": P_COLUMN_FILTER + " ; OwnerFilterField dans toolbar ; API filterOwnerId si sélection",
                "Comportement ancien (owner)": "Liste scopée ; pas de filtre owner",
                "État nouveau front": "API OK, pas RBAC/filtre owner",
                "Filtre owner admin — pattern": P_COLUMN_FILTER,
                "Comportement cible admin": "Afficher toutes les résas par défaut + colonne Propriétaire + filtre en tête (comme demandé)",
                "Comportement cible owner": "Ses réservations seulement",
                "Priorité impl.": "P0",
            },
            {
                "Page / sous-page": "Vue Planning (Gantt réservations)",
                "Route (nouveau)": "/reservations/planning (nav) — route non créée",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Reservation → Séjour (ReservationCalendar)",
                "Comportement ancien (admin)": "requestOwnerId du contexte : null = toutes ; sinon filtre",
                "Comportement ancien (owner)": "Son planning",
                "État nouveau front": "Manquant (lien nav sans route)",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Planning global ; filtre owner optionnel",
                "Comportement cible owner": "Son planning",
                "Priorité impl.": "P1",
            },
            {
                "Page / sous-page": "Détail séjour / réservation",
                "Route (nouveau)": "/reservations/:id",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "admin/Reservation/reservation-detail",
                "Comportement ancien (admin)": "Accès cross-owner ; actions admin (ex. ack annulation)",
                "Comportement ancien (owner)": "Détail de ses résas",
                "État nouveau front": "Partiel",
                "Filtre owner admin — pattern": "Badge owner en lecture seule pour admin",
                "Comportement cible admin": "Voir owner de la résa ; pas de filtre",
                "Comportement cible owner": "Vue normale",
                "Priorité impl.": "P1",
            },
            {
                "Page / sous-page": "Demandes (leads)",
                "Route (nouveau)": "/requests (Service client) — ancien tab Demandes",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Reservation → Demandes (leadReservation)",
                "Comportement ancien (admin)": P_COLUMN_FILTER + " (getAdminOwnerFilterParams)",
                "Comportement ancien (owner)": "Ses demandes",
                "État nouveau front": "Mock / séparé",
                "Filtre owner admin — pattern": P_COLUMN_FILTER,
                "Comportement cible admin": "Liste + filtre owner",
                "Comportement cible owner": "Scope auto",
                "Priorité impl.": "P2",
            },
        ],
    },
    "05_Taches_Operations.xlsx": {
        "tab": "Tâches",
        "rows": [
            {
                "Page / sous-page": "Liste tâches",
                "Route (nouveau)": "/tasks, /tasks/list",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Tasks → Liste (TasksNew)",
                "Comportement ancien (admin)": "AdminOwnerScopeLayout + requestOwnerId : null = toutes tâches",
                "Comportement ancien (owner)": "Tâches de son ownerId",
                "État nouveau front": "Pas de filtre owner",
                "Filtre owner admin — pattern": P_ALL_DEFAULT + " ; colonne owner si liste dense",
                "Comportement cible admin": "Toutes tâches par défaut ; filtre owner ; création tâche avec pick listing cross-owner",
                "Comportement cible owner": "Ses tâches",
                "Priorité impl.": "P0",
            },
            {
                "Page / sous-page": "Équipe",
                "Route (nouveau)": "/tasks/team",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Tasks → Équipe (StaffViewPage)",
                "Comportement ancien (admin)": "requestOwnerId pour charger staff du owner sélectionné",
                "Comportement ancien (owner)": "Son équipe",
                "État nouveau front": "Mock",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Voir tous staff ou filtrer par owner",
                "Comportement cible owner": "Son équipe",
                "Priorité impl.": "P1",
            },
            {
                "Page / sous-page": "Vue Séjour (planning tâches)",
                "Route (nouveau)": "/tasks/planning",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Tasks → Calendrier (UltimateDashboard)",
                "Comportement ancien (admin)": "taskOwnerId via AdminOwnerFilter",
                "Comportement ancien (owner)": "Son calendrier ops",
                "État nouveau front": "Mock",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Calendrier global ; filtre owner",
                "Comportement cible owner": "Auto-scope",
                "Priorité impl.": "P1",
            },
            {
                "Page / sous-page": "Staff WhatsApp",
                "Route (nouveau)": "/tasks/staff-whatsapp",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Tasks → Staff WhatsApp (StaffWAChatPage)",
                "Comportement ancien (admin)": "AdminOwnerScopeLayout ; isAdmin flags dans thread",
                "Comportement ancien (owner)": "Conversations de son équipe",
                "État nouveau front": "Mock",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Tous threads ou filtre owner",
                "Comportement cible owner": "Scope équipe",
                "Priorité impl.": "P2",
            },
            {
                "Page / sous-page": "Task Config (admin)",
                "Route (nouveau)": "Non exposé dans nouveau nav",
                "Accès recommandé": ACCESS_ADMIN,
                "Équivalent ancien dashboard": "admin/TaskConfig, admin/concierge-type",
                "Comportement ancien (admin)": "Admin only Scope",
                "Comportement ancien (owner)": "Pas d'accès menu",
                "État nouveau front": "Absent",
                "Filtre owner admin — pattern": P_BLOCK_OWNER + " (templates par owner)",
                "Comportement cible admin": "Admin only — reproduire dans Settings ou Ops admin",
                "Comportement cible owner": "N/A",
                "Priorité impl.": "P3",
            },
        ],
    },
    "06_Communications.xlsx": {
        "tab": "Communications",
        "rows": [
            {
                "Page / sous-page": "Communications Hub",
                "Route (nouveau)": "/communications/whatsapp (CommsPage)",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Communications → Unified / WhatsApp",
                "Comportement ancien (admin)": "AdminOwnerScopeLayout ; filterOwnerId sur tabs",
                "Comportement ancien (owner)": "Ses conversations",
                "État nouveau front": "Mock",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Inbox global ; filtre owner",
                "Comportement cible owner": "Ses guests",
                "Priorité impl.": "P1",
            },
            {
                "Page / sous-page": "WhatsApp Guests",
                "Route (nouveau)": "/communications/whatsapp-guests",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Communications → WhatsApp",
                "Comportement ancien (admin)": "filterOwnerId via useAdminOwnerFilter",
                "Comportement ancien (owner)": "Scope JWT",
                "État nouveau front": "Partiel",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Tous threads ; filtre owner",
                "Comportement cible owner": "Auto-scope",
                "Priorité impl.": "P1",
            },
            {
                "Page / sous-page": "Staff WhatsApp",
                "Route (nouveau)": "/communications/whatsapp-staff, /communications/staff",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Communications → Staff WhatsApp",
                "Comportement ancien (admin)": "Idem hub",
                "Comportement ancien (owner)": "Son staff",
                "État nouveau front": "Mock",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Filtre owner",
                "Comportement cible owner": "Auto-scope",
                "Priorité impl.": "P2",
            },
            {
                "Page / sous-page": "Templates (QA)",
                "Route (nouveau)": "comms/templates (hub tab)",
                "Accès recommandé": ACCESS_ADMIN,
                "Équivalent ancien dashboard": "Communications → WA templates (QA)",
                "Comportement ancien (admin)": "Scope Admin/SuperAdmin only",
                "Comportement ancien (owner)": "Pas dans menu",
                "État nouveau front": "Mock",
                "Filtre owner admin — pattern": "N/A",
                "Comportement cible admin": "Admin only QA tools",
                "Comportement cible owner": "N/A",
                "Priorité impl.": "P3",
            },
            {
                "Page / sous-page": "Messages OTA",
                "Route (nouveau)": "/communications/messages-ota, /communications/ota",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Communications → OTA Messages",
                "Comportement ancien (admin)": "filterOwnerId",
                "Comportement ancien (owner)": "Scope JWT",
                "État nouveau front": "Mock",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Inbox OTA global + filtre",
                "Comportement cible owner": "Ses messages",
                "Priorité impl.": "P2",
            },
            {
                "Page / sous-page": "Demande / Leads",
                "Route (nouveau)": "comms/leads tab",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Communications → Leads",
                "Comportement ancien (admin)": "filterOwnerId",
                "Comportement ancien (owner)": "Ses leads",
                "État nouveau front": "Mock",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Filtre owner",
                "Comportement cible owner": "Auto-scope",
                "Priorité impl.": "P2",
            },
            {
                "Page / sous-page": "Avis (hub)",
                "Route (nouveau)": "comms/reviews tab",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Communications → Reviews",
                "Comportement ancien (admin)": "filterOwnerId",
                "Comportement ancien (owner)": "Ses avis",
                "État nouveau front": "Mock",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Filtre owner",
                "Comportement cible owner": "Auto-scope",
                "Priorité impl.": "P2",
            },
        ],
    },
    "07_Service_Client.xlsx": {
        "tab": "Service Client",
        "rows": [
            {
                "Page / sous-page": "Demandes",
                "Route (nouveau)": "/requests",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Reservation → Demandes / comms leads",
                "Comportement ancien (admin)": "Filtre owner sur listes",
                "Comportement ancien (owner)": "Ses demandes",
                "État nouveau front": "Mock",
                "Filtre owner admin — pattern": P_COLUMN_FILTER,
                "Comportement cible admin": "Colonne + filtre owner",
                "Comportement cible owner": "Auto-scope",
                "Priorité impl.": "P2",
            },
            {
                "Page / sous-page": "Avis",
                "Route (nouveau)": "/reviews",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "admin/Reviews (legacy) + Communications Reviews",
                "Comportement ancien (admin)": "filterOwnerId",
                "Comportement ancien (owner)": "Ses reviews",
                "État nouveau front": "Mock",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Filtre owner",
                "Comportement cible owner": "Auto-scope",
                "Priorité impl.": "P2",
            },
        ],
    },
    "08_Catalogue.xlsx": {
        "tab": "Catalogue",
        "rows": [
            {
                "Page / sous-page": "Annonces (liste)",
                "Route (nouveau)": "/listings, /catalogue/listings",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Listings → Active / Inactive",
                "Comportement ancien (admin)": "AdminOwnerScopeLayout + filterOwnerId sur getListings ; colonne owner implicite",
                "Comportement ancien (owner)": "Ses annonces",
                "État nouveau front": "API (ownerId/ownerName types) sans filtre UI",
                "Filtre owner admin — pattern": P_COLUMN_FILTER,
                "Comportement cible admin": "Grille toutes annonces + colonne Owner + filtre",
                "Comportement cible owner": "Ses annonces",
                "Priorité impl.": "P0",
            },
            {
                "Page / sous-page": "Créer / éditer annonce",
                "Route (nouveau)": "/listings/:id",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "admin/Listing/new, edit",
                "Comportement ancien (admin)": "Champ ownerId visible (autocomplete owners) ; RU import pick owner",
                "Comportement ancien (owner)": "ownerId hidden = son id",
                "État nouveau front": "Formulaire sans gate owner admin",
                "Filtre owner admin — pattern": "Sélecteur owner obligatoire à la création admin",
                "Comportement cible admin": "Doit choisir owner avant save si création",
                "Comportement cible owner": "ownerId forcé",
                "Priorité impl.": "P1",
            },
            {
                "Page / sous-page": "Tarification",
                "Route (nouveau)": "/pricing, /catalogue/pricing",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Dynamic Pricing → ByListing",
                "Comportement ancien (admin)": "AdminOwnerScopeLayout sur page pricing",
                "Comportement ancien (owner)": "Ses listings",
                "État nouveau front": "Partiel",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Tous listings ; filtre owner",
                "Comportement cible owner": "Auto-scope",
                "Priorité impl.": "P1",
            },
            {
                "Page / sous-page": "Canaux",
                "Route (nouveau)": "/channels, /catalogue/channels",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Channel Manager",
                "Comportement ancien (admin)": "Accès tous owners (JWT)",
                "Comportement ancien (owner)": "Ses connexions",
                "État nouveau front": "Partiel",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Filtre owner sur mapping",
                "Comportement cible owner": "Auto-scope",
                "Priorité impl.": "P2",
            },
            {
                "Page / sous-page": "Clients (CRM léger)",
                "Route (nouveau)": "/clients",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Guests → Direct Guests",
                "Comportement ancien (admin)": "AdminOwnerScopeLayout sur Client.page",
                "Comportement ancien (owner)": "Ses clients",
                "État nouveau front": "Mock",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Filtre owner",
                "Comportement cible owner": "Auto-scope",
                "Priorité impl.": "P2",
            },
            {
                "Page / sous-page": "Contacts WhatsApp",
                "Route (nouveau)": "/clients/contacts",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "Guests → WhatsApp Contacts",
                "Comportement ancien (admin)": "ownerIdFilter dans PublicClientWhiteListGrouped",
                "Comportement ancien (owner)": "Ses contacts",
                "État nouveau front": "Filtre owner mock local (ownerName chips) — pas branché RBAC",
                "Filtre owner admin — pattern": P_COLUMN_FILTER,
                "Comportement cible admin": "Colonne owner + filtre API",
                "Comportement cible owner": "Auto-scope",
                "Priorité impl.": "P2",
            },
            {
                "Page / sous-page": "CRM",
                "Route (nouveau)": "/crm",
                "Accès recommandé": ACCESS_ALL,
                "Équivalent ancien dashboard": "CrmHub (si activé)",
                "Comportement ancien (admin)": "Selon module",
                "Comportement ancien (owner)": "Selon grants",
                "État nouveau front": "Partiel",
                "Filtre owner admin — pattern": P_ALL_DEFAULT,
                "Comportement cible admin": "Filtre owner",
                "Comportement cible owner": "Auto-scope",
                "Priorité impl.": "P3",
            },
            {
                "Page / sous-page": "Onboarding",
                "Route (nouveau)": "/onboarding",
                "Accès recommandé": ACCESS_ADMIN_OWNER,
                "Équivalent ancien dashboard": "(flux inscription owner)",
                "Comportement ancien (admin)": "Peut gérer comptes owners (Team → Property manager)",
                "Comportement ancien (owner)": "Son onboarding",
                "État nouveau front": "Générique",
                "Filtre owner admin — pattern": "N/A",
                "Comportement cible admin": "Création compte owner",
                "Comportement cible owner": "Wizard self-service",
                "Priorité impl.": "P3",
            },
        ],
    },
    "00_RESUME_GLOBAL.xlsx": {
        "tab": "Résumé",
        "rows": [],
    },
}

# Build summary from all rows
summary_rows = []
for fname, spec in SHEETS.items():
    if fname.startswith("00_"):
        continue
    chap = spec["tab"]
    for row in spec["rows"]:
        summary_rows.append(
            {
                "Chapitre nav": chap,
                "Page / sous-page": row["Page / sous-page"],
                "Route (nouveau)": row["Route (nouveau)"],
                "Accès recommandé": row["Accès recommandé"],
                "État nouveau front": row["État nouveau front"],
                "Pattern filtre admin": row["Filtre owner admin — pattern"],
                "Priorité impl.": row["Priorité impl."],
            }
        )

SHEETS["00_RESUME_GLOBAL.xlsx"]["rows"] = summary_rows

# Summary uses slightly different columns
SUMMARY_COLUMNS = [
    ("Chapitre nav", 18),
    ("Page / sous-page", 28),
    ("Route (nouveau)", 32),
    ("Accès recommandé", 22),
    ("État nouveau front", 18),
    ("Pattern filtre admin", 36),
    ("Priorité impl.", 12),
]


def write_summary():
    wb = Workbook()
    ws = wb.active
    ws.title = "Résumé global"
    headers = [h for h, _ in SUMMARY_COLUMNS]
    ws.append(headers)
    for row in summary_rows:
        ws.append([row.get(h, "") for h, _ in SUMMARY_COLUMNS])
    for col_idx, (_, width) in enumerate(SUMMARY_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    path = OUT_DIR / "00_RESUME_GLOBAL.xlsx"
    wb.save(path)
    return path


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    paths = []
    for fname, spec in SHEETS.items():
        if fname.startswith("00_"):
            continue
        paths.append(write_workbook(fname, spec["tab"], spec["rows"]))
    paths.append(write_summary())

    # Patterns reference sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Patterns filtre"
    ws.append(["Pattern", "Description", "Pages typiques"])
    data = [
        ("A — Tous par défaut", P_ALL_DEFAULT, "Tasks, Comms, Calendar, Analytics, Plans"),
        ("B — Colonne + filtre", P_COLUMN_FILTER, "Reservations liste, Listings, Reports history"),
        ("C — Bloquer sans owner", P_BLOCK_OWNER, "Orchestration /config, Settings scoped tabs"),
        ("D — JWT seul", P_NONE, "Overview dashboard (filtre optionnel)"),
        ("E — Chips Admin/Owner", P_ORCH_CHIPS, "Orchestration configuration"),
    ]
    for row in data:
        ws.append(row)
    for col_idx, w in enumerate([22, 48, 36], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    p = OUT_DIR / "99_PATTERNS_REFERENCE.xlsx"
    wb.save(p)
    paths.append(p)

    print("Generated:")
    for p in paths:
        print(" ", p)


if __name__ == "__main__":
    main()
